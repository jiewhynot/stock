from openpyxl import Workbook
wb = Workbook()
ws = wb.active # 激活工作表
ws['A1'] = "股票代码" # A1表格输入数据
ws["B1"]="股票价格"
ws.append(['科比', '1997年', '后卫', '赛季报销']) # 添加一行数据
ws.append(['60000', '6.89'])
wb.save(r'C:\Users\jie1994\Desktop\yueshi\stock.xlsx') # 保存文件